from openpyxl import load_workbook
import click


def save_banking_data(db, xlsx_path):
    wb = load_workbook(filename=xlsx_path, read_only=True)

    # use the first worksheet
    ws = wb[wb.sheetnames[0]]

    translate_fieldnames = {
        'Bezeichnung': 'name',
        'Bank-leitzahl': 'blz',
        'BIC': 'bic',
        'PLZ': 'zipcode',
        'Ort': 'city',
        'Kurzbezeichnung': 'short_description',
        'PAN': 'pan',
        'Prüfziffer-berechnungs-methode': 'check_calculation_method',
        'Datensatz-nummer': 'dataset_number',
        'Merkmal': 'merkmal',
        'Änderungs-kennzeichen': 'change_type',
        'Bankleitzahl-löschung': 'is_deletion',
        'Nachfolge-Bankleitzahl': 'following_blz',
    }

    # translate tablenames to English names
    field_names = []
    for item in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]:
        field_names.append(translate_fieldnames[item])

    num_rows = ws.calculate_dimension().split(":")[1][1:]

    created = "bundesbank_blz" not in db.table_names()

    with click.progressbar(
        ws.iter_rows(min_row=2, values_only=True), length=int(num_rows)
    ) as progress:
        for row in progress:
            db["bundesbank_blz"].insert(
                record=dict(zip(field_names, row)), pk="blz", replace=True)

    if created:
        db["bundesbank_blz"].create_index(["blz"])
        db["bundesbank_blz"].create_index(["bic"])
